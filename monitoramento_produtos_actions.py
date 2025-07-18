import os
import json
import time
import datetime
import base64
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Configurações do Telegram
# Estas serão substituídas pelos secrets do GitHub Actions
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "7538392371:AAH3-eZcq7wrf3Uycv9zPq1PjlSvWfLtYlc")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "-1002593932783")

# Configurações do GitHub
# Estas serão substituídas pelos secrets do GitHub Actions
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPOSITORY = os.environ.get("GITHUB_REPOSITORY", "")
GITHUB_ACTOR = os.environ.get("GITHUB_ACTOR", "")

# Função para obter o horário atual no fuso horário de Brasília (UTC-3)

def salvar_produtos_on(dados_produtos):
    """Salva os produtos que estavam ON na execução atual"""
    produtos_on = [f"{p['Seção']}|{p['Produto']}" for p in dados_produtos if p['Status'] == 'ON']
    with open("produtos_on_ultima_execucao.json", "w", encoding="utf-8") as f:
        json.dump(produtos_on, f, indent=2, ensure_ascii=False)
    fazer_upload_github("produtos_on_ultima_execucao.json", "produtos_on_ultima_execucao.json")

def carregar_produtos_on_anterior():
    """Carrega os produtos que estavam ON na última execução"""
    arquivo = "produtos_on_ultima_execucao.json"
    baixar_arquivo_github(arquivo)
    if not os.path.exists(arquivo):
        return []
    try:
        with open(arquivo, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []

def horario_brasil():
    """Retorna o horário atual no fuso horário de Brasília (UTC-3)"""
    return datetime.datetime.now() - datetime.timedelta(hours=3)

def limpar_preco(texto):
    """Limpa e formata o texto do preço, removendo repetições"""
    if not texto:
        return None
    
    if "R$" in texto:
        partes = texto.split("R$")
        if len(partes) > 1:
            prefixo = partes[0].strip() + " " if partes[0].strip() else ""
            valor = "R$" + partes[1].split()[0].strip()
            return prefixo + valor
    
    return texto.strip()

def extrair_preco(product):
    """Extrai e formata o preço do produto sem repetições"""
    try:
        try:
            price_discount = product.find_element(By.CLASS_NAME, "dish-card__price--discount").text.strip()
            price_discount = limpar_preco(price_discount)
        except NoSuchElementException:
            price_discount = None

        try:
            price_original = product.find_element(By.CLASS_NAME, "dish-card__price--original").text.strip()
            price_original = limpar_preco(price_original)
        except NoSuchElementException:
            price_original = None

        try:
            price_normal = product.find_element(By.CLASS_NAME, "dish-card__price").text.strip()
            price_normal = limpar_preco(price_normal)
        except NoSuchElementException:
            price_normal = None

        if price_discount and price_original:
            return f"De {price_original} por {price_discount}"
        elif price_discount:
            return price_discount
        elif price_original:
            return price_original
        elif price_normal:
            return price_normal
        else:
            return "Preço não encontrado"

    except Exception as e:
        print(f"Erro ao extrair preço: {str(e)}")
        return "Erro ao obter preço"

def salvar_estado_produtos(dados_produtos):
    """Salva o estado atual dos produtos para comparação futura"""
    # No GitHub Actions, salvamos no diretório de trabalho
    arquivo_estado = "estado_produtos.json"
    
    # Criar dicionário com informações essenciais
    estado = {}
    for produto in dados_produtos:
        # Usar nome do produto como chave
        chave = f"{produto['Seção']}|{produto['Produto']}"
        estado[chave] = {
            "Preço": produto["Preço"],
            "Descrição": produto.get("Descrição", ""),
            "Status": produto.get("Status", "ON"),
            "Última verificação": horario_brasil().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    # Salvar no arquivo
    with open(arquivo_estado, "w", encoding="utf-8") as f:
        json.dump(estado, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Estado atual salvo com {len(estado)} produtos")
    
    # Fazer upload do arquivo para o GitHub
    fazer_upload_github(arquivo_estado, arquivo_estado)
    
    return estado

def carregar_estado_anterior():
    """Carrega o estado anterior dos produtos para comparação"""
    arquivo_estado = "estado_produtos.json"
    
    # Tentar baixar o arquivo do GitHub primeiro
    baixar_arquivo_github(arquivo_estado)
    
    if not os.path.exists(arquivo_estado):
        print("⚠️ Nenhum estado anterior encontrado. Esta parece ser a primeira execução.")
        return {}
    
    try:
        with open(arquivo_estado, "r", encoding="utf-8") as f:
            estado = json.load(f)
            print(f"✅ Estado anterior carregado com {len(estado)} produtos")
            return estado
    except Exception as e:
        print(f"❌ Erro ao carregar estado anterior: {str(e)}")
        return {}

    """Carrega o histórico de status dos produtos"""
    arquivo_historico = ""
    
    # Tentar baixar o arquivo do GitHub primeiro
    baixar_arquivo_github(arquivo_historico)
    
    if not os.path.exists(arquivo_historico):
        print("⚠️ Nenhum histórico encontrado. Criando novo arquivo de histórico.")
        return {}
    
    try:
        with open(arquivo_historico, "r", encoding="utf-8") as f:
            historico = json.load(f)
            print(f"✅ Histórico carregado com {len(historico)} produtos")
            return historico
    except Exception as e:
        print(f"❌ Erro ao carregar histórico: {str(e)}")
        return {}

def (dados_produtos, ):
    """Atualiza o histórico de status dos produtos"""
    arquivo_historico = ""
    historico = ()
    
    timestamp = horario_brasil().strftime("%Y-%m-%d %H:%M:%S")
    
    # Atualizar produtos atuais
    for produto in dados_produtos:
        chave = f"{produto['Seção']}|{produto['Produto']}"
        if chave not in historico:
            historico[chave] = {
                "nome": produto["Produto"],
                "secao": produto["Seção"],
                "status_atual": produto["Status"],
                "preco_atual": produto["Preço"],
                "ultima_verificacao": timestamp,
                "historico": []
            }
        else:
            # Se o status mudou, adicionar ao histórico
            if historico[chave]["status_atual"] != produto["Status"]:
                historico[chave]["historico"].append({
                    "status": historico[chave]["status_atual"],
                    "preco": historico[chave]["preco_atual"],
                    "timestamp": historico[chave]["ultima_verificacao"]
                })
            
            # Atualizar status atual
            historico[chave]["status_atual"] = produto["Status"]
            historico[chave]["preco_atual"] = produto["Preço"]
            historico[chave]["ultima_verificacao"] = timestamp
    
    # Atualizar produtos desaparecidos
    for produto in :
        chave = f"{produto['Seção']}|{produto['Produto']}"
        if chave not in historico:
            historico[chave] = {
                "nome": produto["Produto"],
                "secao": produto["Seção"],
                "status_atual": "",
                "preco_atual": produto["Preço"],
                "ultima_verificacao": timestamp,
                "historico": []
            }
        else:
            # Se o status mudou, adicionar ao histórico
            if historico[chave]["status_atual"] != "":
                historico[chave]["historico"].append({
                    "status": historico[chave]["status_atual"],
                    "preco": historico[chave]["preco_atual"],
                    "timestamp": historico[chave]["ultima_verificacao"]
                })
            
            # Atualizar status atual
            historico[chave]["status_atual"] = ""
            historico[chave]["ultima_verificacao"] = timestamp
    
    # Salvar histórico atualizado
    with open(arquivo_historico, "w", encoding="utf-8") as f:
        json.dump(historico, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Histórico atualizado com {len(historico)} produtos")
    
    # Fazer upload do arquivo para o GitHub
    fazer_upload_github(arquivo_historico, arquivo_historico)
    
    return historico

def salvar_produtos_on_atual(produtos):
    produtos_on = [
        f"{p['Seção']}|{p['Produto']}"
        for p in produtos
        if p.get("Status") == "ON"
    ]
    with open("produtos_on_ultima_execucao.json", "w", encoding="utf-8") as f:
        json.dump(produtos_on, f, ensure_ascii=False, indent=2)
